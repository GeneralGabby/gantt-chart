import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_PURPLE_FILL  = PatternFill("solid", fgColor="4B3F72")  # % Complete (completed)
MED_PURPLE_FILL   = PatternFill("solid", fgColor="7B6EA6")  # Plan Duration (in progress)
LIGHT_PURPLE_FILL = PatternFill("solid", fgColor="C8B8E8")  # Plan only (remaining)
GOLD_FILL         = PatternFill("solid", fgColor="F0A500")  # Actual beyond plan
HEADER_FILL       = PatternFill("solid", fgColor="4B3F72")  # Header background
SUBHDR_FILL       = PatternFill("solid", fgColor="7B6EA6")  # Alternating period header
ALT_ROW_FILL      = PatternFill("solid", fgColor="F5F0FF")  # Alternating row background
WHITE_FILL        = PatternFill("solid", fgColor="FFFFFF")
PERIOD_ALT_FILL   = PatternFill("solid", fgColor="EDE8F5")  # Light period bg

# ── Borders ───────────────────────────────────────────────────────────────────
THIN_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE
)

# ── Fonts ─────────────────────────────────────────────────────────────────────
WHITE_BOLD_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=10)
DARK_BOLD_FONT   = Font(name="Calibri", bold=True,  color="4B3F72", size=10)
NORM_FONT        = Font(name="Calibri", bold=False, color="333333", size=9)
TITLE_FONT       = Font(name="Calibri", bold=True,  color="4B3F72", size=18)
PERIOD_HDR_FONT  = Font(name="Calibri", bold=True,  color="FFFFFF", size=8)
LEGEND_FONT      = Font(name="Calibri", size=8, color="333333")

# ── Alignments ────────────────────────────────────────────────────────────────
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Task data ─────────────────────────────────────────────────────────────────
tasks = [
    (1,  "Project Initiation",                                       1,  5,  1,  5,  100, "Project Leader (Kemi)"),
    (2,  "Client introduction & first meeting",                      2,  1,  2,  1,  100, "Proj. Leader & Comms Lead (Sonia)"),
    (3,  "Assign roles to team members",                             3,  1,  3,  1,  100, "Project Leader (Kemi)"),
    (4,  "Trello setup & add group members",                         3,  1,  3,  1,  100, "Deputy / Coordinator (Gabriel)"),
    (5,  "Compose introductory emails / client comms",               3,  2,  3,  2,  100, "Deputy / Coordinator (Gabriel)"),
    (6,  "Internal group meeting (Gabriel covering Kemi)",           4,  1,  4,  1,  100, "Deputy / Coordinator (Gabriel)"),
    (7,  "Desk-based research on policies & frameworks",             4,  10, 4,  10, 50,  "Research Lead (Grace) & Teams"),
    (8,  "Share research topics & assign teams",                     6,  1,  6,  1,  100, "Research Lead (Grace)"),
    (9,  "Collect case studies (HE & NHS)",                          6,  5,  6,  5,  50,  "Teams 1,2,3"),
    (10, "Summarize known outcomes & results",                       11, 5,  11, 5,  0,   "Data & Analysis Lead (Razvan)"),
    (11, "Compare energy & water-saving measures",                   16, 5,  16, 5,  0,   "Data Lead & Research Teams"),
    (12, "Identify barriers & differences across institutions",      21, 5,  21, 5,  0,   "Data Lead & Research Teams"),
    (13, "Draft report sections",                                    26, 5,  26, 5,  0,   "Documentation Lead (Osas)"),
    (14, "Integrate team contributions",                             31, 5,  31, 5,  0,   "Documentation Lead (Osas)"),
    (15, "Proofread, format, and check accuracy",                    36, 5,  36, 5,  0,   "QA Lead (Chathu)"),
    (16, "Prepare slides for client / internal follow-up",           36, 3,  36, 3,  100, "QA & Presentation Lead (Chathu)"),
    (17, "Prepare final slides & presentation",                      41, 5,  41, 5,  0,   "QA & Presentation Lead (Chathu)"),
    (18, "Submission & Final Delivery",                              46, 1,  46, 1,  0,   "Proj. Leader & Documentation Lead"),
    (19, "Take meeting minutes (Sonia)",                             2,  19, 2,  19, 50,  "Communications Lead (Sonia)"),
    (20, "Facilitate meetings with client (Kemi)",                   2,  19, 2,  19, 50,  "Project Leader (Kemi)"),
]

MAX_PERIOD = 47

# ── Column indices ────────────────────────────────────────────────────────────
COL_TASKID   = 1   # A
COL_ACTIVITY = 2   # B
COL_PSTART   = 3   # C
COL_PDUR     = 4   # D
COL_ASTART   = 5   # E
COL_ADUR     = 6   # F
COL_PCT      = 7   # G
COL_RESP     = 8   # H
COL_PERIOD1  = 9   # I – period 1

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(COL_TASKID)].width   = 7
ws.column_dimensions[get_column_letter(COL_ACTIVITY)].width = 38
ws.column_dimensions[get_column_letter(COL_PSTART)].width   = 11
ws.column_dimensions[get_column_letter(COL_PDUR)].width     = 11
ws.column_dimensions[get_column_letter(COL_ASTART)].width   = 11
ws.column_dimensions[get_column_letter(COL_ADUR)].width     = 11
ws.column_dimensions[get_column_letter(COL_PCT)].width      = 10
ws.column_dimensions[get_column_letter(COL_RESP)].width     = 28
for p in range(MAX_PERIOD):
    ws.column_dimensions[get_column_letter(COL_PERIOD1 + p)].width = 2.5

# ── Row heights ───────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36  # title
ws.row_dimensions[2].height = 18  # legend
ws.row_dimensions[3].height = 32  # column headers
for r in range(4, 4 + len(tasks)):
    ws.row_dimensions[r].height = 18

# ── Row 1 – Title ─────────────────────────────────────────────────────────────
last_col = COL_PERIOD1 + MAX_PERIOD - 1
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
title_cell = ws.cell(row=1, column=1, value="PROJECT PLANNER \u2013 GANTT CHART")
title_cell.font      = TITLE_FONT
title_cell.alignment = LEFT_ALIGN
title_cell.fill      = WHITE_FILL

# ── Row 2 – Legend ────────────────────────────────────────────────────────────
ws.cell(row=2, column=1, value="Legend:").font      = DARK_BOLD_FONT
ws.cell(row=2, column=1).alignment                  = LEFT_ALIGN

legend_items = [
    (COL_PERIOD1,      DARK_PURPLE_FILL,  "% Complete (done)"),
    (COL_PERIOD1 + 5,  MED_PURPLE_FILL,   "Plan Duration (in progress)"),
    (COL_PERIOD1 + 11, LIGHT_PURPLE_FILL, "Plan only (remaining)"),
    (COL_PERIOD1 + 17, GOLD_FILL,         "Actual beyond plan"),
]
for swatch_col, fill, label in legend_items:
    swatch = ws.cell(row=2, column=swatch_col)
    swatch.fill      = fill
    swatch.border    = THIN_BORDER
    txt = ws.cell(row=2, column=swatch_col + 1, value=label)
    txt.font         = LEGEND_FONT
    txt.alignment    = LEFT_ALIGN

# ── Row 3 – Column headers ────────────────────────────────────────────────────
col_headers = [
    "TASK ID", "ACTIVITY", "PLAN\nSTART", "PLAN\nDURATION",
    "ACTUAL\nSTART", "ACTUAL\nDURATION", "%\nCOMPLETE", "RESPONSIBLE LEAD / TEAM",
]
for idx, header in enumerate(col_headers, start=1):
    cell = ws.cell(row=3, column=idx, value=header)
    cell.fill      = HEADER_FILL
    cell.font      = WHITE_BOLD_FONT
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER

for p in range(1, MAX_PERIOD + 1):
    col = COL_PERIOD1 + p - 1
    cell = ws.cell(row=3, column=col, value=str(p))
    cell.fill      = HEADER_FILL if p % 2 == 1 else SUBHDR_FILL
    cell.font      = PERIOD_HDR_FONT
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER

# ── Helper to write a styled data cell ───────────────────────────────────────
def write_cell(row, col, value, fill, font, align):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill
    c.font      = font
    c.alignment = align
    c.border    = THIN_BORDER
    return c

# ── Rows 4+ – Task data & Gantt bars ─────────────────────────────────────────
for i, (tid, activity, ps, pd, as_, ad, pct, resp) in enumerate(tasks):
    row       = 4 + i
    row_fill  = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL

    write_cell(row, COL_TASKID,   tid,        row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_ACTIVITY, activity,   row_fill, NORM_FONT, LEFT_ALIGN)
    write_cell(row, COL_PSTART,   ps,         row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_PDUR,     pd,         row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_ASTART,   as_,        row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_ADUR,     ad,         row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_PCT,      f"{pct}%",  row_fill, NORM_FONT, CENTER_ALIGN)
    write_cell(row, COL_RESP,     resp,       row_fill, NORM_FONT, LEFT_ALIGN)

    # Derived period boundaries
    plan_end     = ps + pd - 1
    actual_end   = as_ + ad - 1
    # Number of completed periods within the plan (rounded to nearest whole period)
    completed_periods = round(pd * pct / 100)
    complete_end = ps + completed_periods - 1 if pct > 0 else ps - 1

    for p in range(1, MAX_PERIOD + 1):
        col      = COL_PERIOD1 + p - 1
        in_plan   = ps  <= p <= plan_end
        in_actual = as_ <= p <= actual_end
        in_done   = (pct > 0) and (as_ <= p <= complete_end)

        cell        = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER

        if in_plan and in_done:
            # Completed portion within plan → dark purple
            cell.fill = DARK_PURPLE_FILL
        elif in_plan and in_actual:
            # Planned & actually in progress (not yet complete) → medium purple
            cell.fill = MED_PURPLE_FILL
        elif in_plan:
            # Plan only, not yet started → light purple
            cell.fill = LIGHT_PURPLE_FILL
        elif in_actual:
            # Actual work beyond the plan → gold
            cell.fill = GOLD_FILL
        else:
            # No bar – alternating background
            cell.fill = PERIOD_ALT_FILL if p % 2 == 0 else WHITE_FILL

# ── Freeze panes at I4 (column I = period 1, row 4 = first data row) ─────────
ws.freeze_panes = "I4"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save("Gantt_Chart.xlsx")
print("\u2705 Gantt_Chart.xlsx saved successfully!")
